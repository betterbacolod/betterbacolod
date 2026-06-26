#!/usr/bin/env python3
"""
Convert DOE electric grid XLSX data into src/data/energy/electric-grid.json.

Usage:
    python3 scripts/import-doe-electric-grid.py <path-to-xlsx>

The workbook must include:
    f_Demand     Year | Island Group | Demand | Start of Year
    f_Generation FACILITY NAME (DOE Coding) | ... | REGION | ...
"""
import json
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl

OUTPUT_PATH = Path('src/data/energy/electric-grid.json')
SOURCE_POWER_STATS_URL = 'https://doe.gov.ph/articles/3512596--2025-power-statistics'
SOURCE_PLANTS_URL = (
    'https://doe.gov.ph/articles/3501260--list-of-existing-power-plants-grid-connected-as-of-may-2026'
)
SYSTEM_DEMAND_DOC_URL = (
    'https://doe.gov.ph/documents/d/guest/annex-6_system-peak-demand-per-grid-2001-2025-pdf'
)
VISAYAS_PLANTS_DOC_URL = (
    'https://doe.gov.ph/documents/d/guest/02_doe-epimb_loepp-as-of-may-2026_lvm-grid-clean-visayas-pdf'
)


def normalize_header(value):
    return str(value or '').strip()


def iso_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def to_float(value):
    if value is None or value == '':
        return 0.0
    return float(value)


def round_mw(value):
    return round(float(value), 1)


def rows_by_header(ws):
    header_row = next(ws.iter_rows(values_only=True))
    headers = [normalize_header(h) for h in header_row]
    for row in ws.iter_rows(min_row=2, values_only=True):
        yield {headers[i]: row[i] for i in range(len(headers))}


def main(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    demand_rows = []
    for row in rows_by_header(wb['f_Demand']):
        if row['Island Group'] != 'Visayas':
            continue
        demand_rows.append({
            'year': int(row['Year']),
            'demandMw': round_mw(row['Demand']),
            'startOfYear': iso_date(row['Start of Year']),
        })
    demand_rows.sort(key=lambda r: r['year'])

    facilities = []
    for row in rows_by_header(wb['f_Generation']):
        if normalize_header(row['REGION']) != 'NIR':
            continue
        facilities.append({
            'facilityName': normalize_header(row['FACILITY NAME (DOE Coding)']),
            'officialName': normalize_header(
                row['Official Facility Name (as per ERC COC/PAO)'],
            ),
            'resourceType': normalize_header(row['RESOURCE TYPE']),
            'technologyType': normalize_header(row['TECHNOLOGY TYPE']),
            'connectionType': normalize_header(row['TYPE OF CONNECTION']),
            'installedMw': round_mw(row['INSTALLED']),
            'dependableMw': round_mw(row['DEPENDABLE']),
            'units': int(row['NUMBER OF UNITS']) if row['NUMBER OF UNITS'] else None,
            'province': normalize_header(row['PROVINCE']),
            'cityMunicipality': normalize_header(row['CITY/ MUNICIPALITY']),
            'municipalityProvince': normalize_header(row['MUNICIPALITY/ PROVINCE']),
            'operator': normalize_header(row['OPERATOR']),
            'ownerIppa': normalize_header(row['OWNER / IPPA']),
            'commissioned': normalize_header(
                row['DATE COMMISSIONED / COMMERCIAL OPERATION'],
            ),
            'latestYear': int(row['Latest Year']),
            'startOfYear': iso_date(row['Start of Year']),
        })
    facilities.sort(key=lambda r: (-r['installedMw'], r['facilityName']))

    by_resource = aggregate(facilities, 'resourceType')
    by_province = aggregate(facilities, 'province')
    demand_latest = demand_rows[-1]
    demand_first = demand_rows[0]
    demand_growth = (
        (demand_latest['demandMw'] - demand_first['demandMw'])
        / demand_first['demandMw']
        * 100
    )
    installed_total = sum(f['installedMw'] for f in facilities)
    dependable_total = sum(f['dependableMw'] for f in facilities)

    out = {
        'source': {
            'name': 'Department of Energy Philippines',
            'workbook': Path(xlsx_path).name,
            'articles': [
                {
                    'title': '2025 Power Statistics',
                    'url': SOURCE_POWER_STATS_URL,
                },
                {
                    'title': 'List of Existing Power Plants (Grid Connected) as of May 2026',
                    'url': SOURCE_PLANTS_URL,
                },
            ],
            'documents': [
                {
                    'title': 'System Peak Demand per Grid, 2001-2025',
                    'url': SYSTEM_DEMAND_DOC_URL,
                },
                {
                    'title': 'Visayas Existing Power Plants as of May 2026',
                    'url': VISAYAS_PLANTS_DOC_URL,
                },
            ],
        },
        'lastUpdated': datetime.now().date().isoformat(),
        'scopeNote': (
            'Visayas demand and NIR generation context only. The dataset does not '
            'identify a Bacolod City generation facility.'
        ),
        'demand': {
            'islandGroup': 'Visayas',
            'unit': 'MW',
            'rows': demand_rows,
        },
        'generation': {
            'region': 'NIR',
            'unit': 'MW',
            'facilities': facilities,
            'byResourceType': by_resource,
            'byProvince': by_province,
        },
        'stats': {
            'visayasDemand2025Mw': demand_latest['demandMw'],
            'visayasDemandGrowthSince2001Pct': round(demand_growth, 1),
            'visayasDemandRows': len(demand_rows),
            'nirFacilities': len(facilities),
            'nirInstalledMw': round_mw(installed_total),
            'nirDependableMw': round_mw(dependable_total),
            'negrosOccidentalInstalledMw': round_mw(
                sum(f['installedMw'] for f in facilities if f['province'] == 'Negros Occidental'),
            ),
        },
    }

    validate(out)
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(out, indent=2) + '\n')
    print(
        f"Wrote {len(demand_rows)} Visayas demand rows and "
        f"{len(facilities)} NIR facilities to {OUTPUT_PATH}",
    )


def aggregate(facilities, key):
    buckets = defaultdict(list)
    for facility in facilities:
        buckets[facility[key]].append(facility)
    return [
        {
            'name': name,
            'facilityCount': len(items),
            'installedMw': round_mw(sum(i['installedMw'] for i in items)),
            'dependableMw': round_mw(sum(i['dependableMw'] for i in items)),
        }
        for name, items in sorted(
            buckets.items(),
            key=lambda item: (-sum(i['installedMw'] for i in item[1]), item[0]),
        )
    ]


def validate(data):
    assert data['stats']['visayasDemandRows'] == 25
    assert data['stats']['nirFacilities'] == 25
    assert data['stats']['nirInstalledMw'] == 1107.9
    assert data['stats']['nirDependableMw'] == 863.9
    for facility in data['generation']['facilities']:
        assert not (
            facility['province'] == 'Negros Occidental'
            and facility['cityMunicipality'].lower() == 'bacolod'
        )


if __name__ == '__main__':
    if len(sys.argv) != 2:
        print('Usage: python3 scripts/import-doe-electric-grid.py <path-to-xlsx>')
        sys.exit(1)
    main(sys.argv[1])
