#!/usr/bin/env python3
"""Build the Bacolod annual-budget runtime dataset from FDP ABR workbooks.

The workbooks use cached spreadsheet values. This importer deliberately reads
only the proposed-budget column and the top-level rows that are comparable
across the published 2022–2025 Annual Budget Reports.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import openpyxl


REPORTS = (
    {
        "year": 2022,
        "filename": "2022_abr-1.xlsx",
        "url": "https://bacolodcity.gov.ph/wp-content/uploads/2022/08/2022_abr-1.xlsx",
    },
    {
        "year": 2023,
        "filename": "2023_abr.xlsx",
        "url": "https://bacolodcity.gov.ph/wp-content/uploads/2023/02/2023_abr.xlsx",
    },
    {
        "year": 2024,
        "filename": "ABR-2024.xlsx",
        "url": "https://bacolodcity.gov.ph/wp-content/uploads/2024/02/ABR-2024.xlsx",
    },
    {
        "year": 2025,
        "filename": "2025_ABR.xlsx",
        "url": "https://bacolodcity.gov.ph/wp-content/uploads/2022/08/2025_ABR.xlsx",
    },
)

SOURCE_PAGE = "https://bacolodcity.gov.ph/full-disclosure-policy/"
PROPOSED_COLUMN = "S"
REQUIRED_ROWS = {
    "taxRevenue": "Total Tax Revenue",
    "nonTaxRevenue": "Total Non-Tax Revenue",
    "externalSources": "Total External Sources",
    "totalReceipts": "TOTAL RECEIPTS",
    "personnelServices": "TOTAL PERSONAL SERVICES",
    "maintenanceAndOperating": "TOTAL MAINTENANCE & OPERATING EXPENSES",
    "financialExpenses": "TOTAL FINANCIAL EXPENSES",
    "propertyPlantAndEquipment": "TOTAL PROPERTY, PLANT & EQUIPMENT",
    "specialPurposeAppropriations": "TOTAL SPECIAL PURPOSE APPROPRIATION",
    "totalExpenditures": "Total Expenditures",
}
EXPENDITURE_START = "Salaries and Wages - Regular Pay"
EXPENDITURE_SECTIONS = (
    ("personnelServices", "Personnel services"),
    ("maintenanceAndOperating", "MOOE"),
    ("financialExpenses", "Financial expenses"),
    ("propertyPlantAndEquipment", "Capital outlay"),
    ("specialPurposeAppropriations", "Special purpose"),
)


def centavos(value: object) -> int:
    if value is None:
        raise ValueError("Missing proposed-budget value")
    return int((Decimal(str(value)) * 100).quantize(Decimal("1"), ROUND_HALF_UP))


def canonical_label(value: object) -> str:
    return " ".join(str(value or "").split()).casefold()


def extract_report(source_dir: Path, report: dict[str, object]) -> dict[str, object]:
    source_path = source_dir / str(report["filename"])
    if not source_path.exists():
        raise FileNotFoundError(source_path)

    workbook = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
    worksheet = workbook.active
    remaining = {key: canonical_label(label) for key, label in REQUIRED_ROWS.items()}
    values: dict[str, int] = {}
    source_rows: dict[str, str] = {}

    for row in worksheet.iter_rows():
        labels = [
            cell
            for cell in row[:14]
            if canonical_label(cell.value) in remaining.values()
        ]
        for label_cell in labels:
            label = canonical_label(label_cell.value)
            key = next(key for key, expected in remaining.items() if expected == label)
            # 2025 contains a second, zero-value "Total External Sources" row.
            # The first matching row is the funding total used by the report.
            values[key] = centavos(
                worksheet[f"{PROPOSED_COLUMN}{label_cell.row}"].value
            )
            source_rows[key] = (
                f"{label_cell.coordinate}:{PROPOSED_COLUMN}{label_cell.row}"
            )
            del remaining[key]

    if remaining:
        raise ValueError(
            f"{source_path.name}: missing required rows: {', '.join(remaining)}"
        )

    local_sources = values["taxRevenue"] + values["nonTaxRevenue"]
    comparable_expenditures = sum(
        values[key]
        for key in (
            "personnelServices",
            "maintenanceAndOperating",
            "financialExpenses",
            "propertyPlantAndEquipment",
            "specialPurposeAppropriations",
        )
    )
    if values["totalReceipts"] != values["totalExpenditures"]:
        raise ValueError(f"{source_path.name}: total receipts do not equal expenditures")
    if comparable_expenditures != values["totalExpenditures"]:
        raise ValueError(
            f"{source_path.name}: expenditure components do not reconcile to total"
        )

    expenditure_start = None
    for row in worksheet.iter_rows():
        label_cell = next(
            (
                cell
                for cell in row[:14]
                if canonical_label(cell.value) == canonical_label(EXPENDITURE_START)
            ),
            None,
        )
        if label_cell is not None:
            expenditure_start = label_cell.row
            break
    if expenditure_start is None:
        raise ValueError(f"{source_path.name}: missing expenditure start row")

    section_ends = [
        (int(source_rows[key].split(":")[0][1:]), category)
        for key, category in EXPENDITURE_SECTIONS
    ]
    expenditure_end = int(source_rows["totalExpenditures"].split(":")[0][1:])
    top_items: list[dict[str, object]] = []
    for row in worksheet.iter_rows(
        min_row=expenditure_start, max_row=expenditure_end - 1
    ):
        row_num = next(cell.row for cell in row if hasattr(cell, "row"))
        label_parts = [
            str(cell.value).strip()
            for cell in row[:14]
            if isinstance(cell.value, str)
            and str(cell.value).strip()
            and any(character.isalpha() for character in str(cell.value))
        ]
        label = " ".join(label_parts)
        amount = worksheet[f"{PROPOSED_COLUMN}{row_num}"].value
        if not label or not isinstance(amount, (int, float)) or amount <= 0:
            continue
        if canonical_label(label).startswith("total "):
            continue
        category = next(
            category for end_row, category in section_ends if row_num < end_row
        )
        top_items.append(
            {
                "label": label,
                "category": category,
                "amountCentavos": centavos(amount),
                "sourceRow": f"S{row_num}",
            }
        )
    top_items.sort(key=lambda item: int(item["amountCentavos"]), reverse=True)

    with source_path.open("rb") as source_file:
        workbook_hash = hashlib.file_digest(source_file, "sha256").hexdigest()

    return {
        "year": report["year"],
        "workbook": report["filename"],
        "workbookSha256": workbook_hash,
        "sheet": worksheet.title,
        "sourceUrl": report["url"],
        "sourceRows": source_rows,
        "proposedAmountCentavos": values["totalExpenditures"],
        "totalReceiptsCentavos": values["totalReceipts"],
        "financing": {
            "localSourcesCentavos": local_sources,
            "externalSourcesCentavos": values["externalSources"],
        },
        "expenditures": {
            "personnelServicesCentavos": values["personnelServices"],
            "maintenanceAndOperatingCentavos": values["maintenanceAndOperating"],
            "financialExpensesCentavos": values["financialExpenses"],
            "propertyPlantAndEquipmentCentavos": values[
                "propertyPlantAndEquipment"
            ],
            "specialPurposeAppropriationsCentavos": values[
                "specialPurposeAppropriations"
            ],
        },
        "topSpendingItems": top_items[:6],
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("source_dir", type=Path)
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("src/data/transparency/bacolod-annual-budget.json"),
    )
    parser.add_argument(
        "--csv-output",
        type=Path,
        default=Path("public/data/bacolod-annual-budget-2022-2025.csv"),
    )
    parser.add_argument("--check", action="store_true")
    args = parser.parse_args()

    payload = {
        "schemaVersion": 1,
        "currency": "PHP",
        "measure": "Proposed annual budget",
        "sourcePage": SOURCE_PAGE,
        "reports": [extract_report(args.source_dir, report) for report in REPORTS],
    }
    encoded = json.dumps(payload, indent=2) + "\n"
    csv_buffer = io.StringIO(newline="")
    writer = csv.writer(csv_buffer)
    writer.writerow(
        [
            "year",
            "measure",
            "proposed_amount_centavos",
            "local_sources_centavos",
            "external_sources_centavos",
            "personnel_services_centavos",
            "maintenance_and_operating_centavos",
            "financial_expenses_centavos",
            "property_plant_and_equipment_centavos",
            "special_purpose_appropriations_centavos",
            "source_url",
        ]
    )
    for report in payload["reports"]:
        financing = report["financing"]
        expenditures = report["expenditures"]
        writer.writerow(
            [
                report["year"],
                payload["measure"],
                report["proposedAmountCentavos"],
                financing["localSourcesCentavos"],
                financing["externalSourcesCentavos"],
                expenditures["personnelServicesCentavos"],
                expenditures["maintenanceAndOperatingCentavos"],
                expenditures["financialExpensesCentavos"],
                expenditures["propertyPlantAndEquipmentCentavos"],
                expenditures["specialPurposeAppropriationsCentavos"],
                report["sourceUrl"],
            ]
        )
    csv_encoded = csv_buffer.getvalue()
    if args.check:
        if not args.output.exists() or args.output.read_text() != encoded:
            raise ValueError(f"{args.output}: generated dataset is out of date")
        if args.csv_output.exists():
            with args.csv_output.open(newline="") as csv_file:
                csv_current = csv_file.read()
        else:
            csv_current = None
        if csv_current != csv_encoded:
            raise ValueError(f"{args.csv_output}: generated CSV is out of date")
        print(f"Validated {args.output} and {args.csv_output}")
        return 0

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(encoded)
    args.csv_output.parent.mkdir(parents=True, exist_ok=True)
    args.csv_output.write_text(csv_encoded)
    print(f"Wrote {args.output} and {args.csv_output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
